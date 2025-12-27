"""
Flask Application for PyPotteryLens
Migrated from Gradio to Flask with native HTML, CSS, and JavaScript
"""

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, Response
from pathlib import Path
import os
import json
import base64
from io import BytesIO
import pandas as pd
import numpy as np
from werkzeug.utils import secure_filename
import torch
import gc
import threading
import time

from utils import (
    PDFProcessor,
    ModelProcessor,
    MaskExtractor,
    AnnotationProcessor,
    ImageProcessor,
    TabularProcessor,
    SecondStepProcessor,
    ExportProcessor,
    MetadataExtractor,
    PDFConfig,
    ModelConfig,
    MaskExtractionConfig,
    AnnotationConfig,
    TabularConfig,
    SecondStepConfig,
    ExportConfig,
    MetadataExtractionConfig
)

from project_manager import ProjectManager
from settings_manager import get_settings_manager
from scale_detector import ScaleBarDetector, ScaleBarConfig
from ai_extractor import get_extractor, image_to_base64, detect_image_media_type, BatchExtractor

app = Flask(__name__)
app.config['SECRET_KEY'] = 'pypotterylens-secret-key-2024'
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = Path('temp_uploads')
app.config['UPLOAD_FOLDER'].mkdir(exist_ok=True)

# Initialize Project Manager
project_manager = ProjectManager(projects_root="projects")

# Initialization status tracking
init_status = {
    'ready': False,
    'stage': 'starting',
    'progress': 0,
    'message': 'Initializing...'
}

def update_init_status(stage, progress, message):
    """Update initialization status"""
    global init_status
    init_status['stage'] = stage
    init_status['progress'] = progress
    init_status['message'] = message
    print(f"[Init] {progress}% - {message}")

# Global progress tracking
operation_progress = {
    'active': False,
    'operation': '',
    'total': 0,
    'current': 0,
    'message': '',
    'percent': 0
}

def update_operation_progress(operation, current, total, message=''):
    """Update operation progress for frontend polling"""
    global operation_progress
    operation_progress['active'] = True
    operation_progress['operation'] = operation
    operation_progress['current'] = current
    operation_progress['total'] = total
    operation_progress['message'] = message
    operation_progress['percent'] = int((current / total * 100)) if total > 0 else 0
    print(f"[{operation}] {operation_progress['percent']}% - {message}")

def clear_operation_progress():
    """Clear operation progress"""
    global operation_progress
    operation_progress['active'] = False
    operation_progress['operation'] = ''
    operation_progress['total'] = 0
    operation_progress['current'] = 0
    operation_progress['message'] = ''
    operation_progress['percent'] = 0

# Initialize directories
ROOT_DIR = Path(".")
PRED_OUTPUT_DIR = ROOT_DIR / "outputs"
PDFIMG_OUTPUT_DIR = ROOT_DIR / "pdf2img_outputs"
MODELS_DIR = ROOT_DIR / "models_vision"
MODELS_CLASSIFIER_DIR = ROOT_DIR / "models_classifier"
ASSETS_DIR = ROOT_DIR / "imgs"

# Create necessary directories
for directory in [PDFIMG_OUTPUT_DIR, MODELS_DIR, PRED_OUTPUT_DIR, MODELS_CLASSIFIER_DIR]:
    directory.mkdir(exist_ok=True)


# ==================== MODEL INITIALIZATION ====================

def download_model(url, destination, model_name, base_progress, progress_range):
    """Download a model file with progress tracking for splash screen"""
    import urllib.request
    import sys
    
    print(f"\n📥 Downloading {model_name}...")
    print(f"   URL: {url}")
    print(f"   Destination: {destination}")
    
    def show_progress(block_num, block_size, total_size):
        downloaded = block_num * block_size
        if total_size > 0:
            percent = min(downloaded * 100 / total_size, 100)
            
            # Update splash screen progress
            current_progress = base_progress + (percent / 100.0) * progress_range
            downloaded_mb = downloaded / (1024 * 1024)
            total_mb = total_size / (1024 * 1024)
            update_init_status('downloading_models', current_progress, 
                             f'Downloading {model_name}: {downloaded_mb:.1f}/{total_mb:.1f} MB')
            
            # Console progress bar
            bar_length = 50
            filled_length = int(bar_length * percent / 100)
            bar = '█' * filled_length + '░' * (bar_length - filled_length)
            
            sys.stdout.write(f'\r   [{bar}] {percent:.1f}% ({downloaded_mb:.1f}/{total_mb:.1f} MB)')
            sys.stdout.flush()
    
    try:
        urllib.request.urlretrieve(url, destination, show_progress)
        print(f"\n   ✅ Successfully downloaded {model_name}")
        return True
    except Exception as e:
        print(f"\n   ❌ Error downloading {model_name}: {e}")
        return False


def initialize_models():
    """Check and download required models if missing"""
    update_init_status('checking_models', 10, 'Checking required models...')
    
    print("\n" + "="*80)
    print(" 🔍 Checking Required Models")
    print("="*80)
    
    models_to_check = {
        'Vision Model': {
            'path': MODELS_DIR / 'BasicModelv8_v01.pt',
            'url': 'https://huggingface.co/lrncrd/PyPotteryLens/resolve/main/BasicModelv8_v01.pt'
        },
        'Classifier Model': {
            'path': MODELS_CLASSIFIER_DIR / 'model_classifier.pth',
            'url': 'https://huggingface.co/lrncrd/PyPotteryLens/resolve/main/model_classifier.pth'
        }
    }
    
    missing_models = []
    
    for model_name, info in models_to_check.items():
        if info['path'].exists():
            print(f"✅ {model_name}: Found at {info['path']}")
        else:
            print(f"❌ {model_name}: Not found at {info['path']}")
            missing_models.append((model_name, info))
    
    if missing_models:
        print(f"\n⚠️  {len(missing_models)} model(s) need to be downloaded")
        print("="*80)
        
        progress_per_model = 60 / len(missing_models)  # 60% total for models (20% -> 80%)
        
        for idx, (model_name, info) in enumerate(missing_models):
            base_progress = 20 + (idx * progress_per_model)
            success = download_model(info['url'], info['path'], model_name, 
                                   base_progress, progress_per_model)
            
            if not success:
                print(f"\n⚠️  Warning: Could not download {model_name}")
                print(f"   Please download manually from: {info['url']}")
                print(f"   And place it at: {info['path']}")
        
        print("\n" + "="*80)
        print(" ✨ Model initialization complete!")
        print("="*80)
    else:
        print("\n✨ All required models are present!")
        print("="*80)
    
    update_init_status('models_ready', 80, 'Models ready, initializing processors...')


def initialize_processors():
    """Initialize all processors after models are ready"""
    global pdf_processor, model_processor, mask_extractor, annotation_processor
    global image_processor, tabular_processor, second_step_processor, export_processor
    
    update_init_status('init_processors', 85, 'Initializing PDF processor...')
    pdf_processor = PDFProcessor(PDFConfig(output_dir=PDFIMG_OUTPUT_DIR))

    update_init_status('init_processors', 88, 'Initializing model processor...')
    model_processor = ModelProcessor(ModelConfig(
        models_dir=MODELS_DIR,
        pred_output_dir=PRED_OUTPUT_DIR
    ))

    update_init_status('init_processors', 90, 'Initializing mask extractor...')
    mask_extractor = MaskExtractor(MaskExtractionConfig(
        pdfimg_output_dir=PDFIMG_OUTPUT_DIR,
        pred_output_dir=PRED_OUTPUT_DIR
    ))

    update_init_status('init_processors', 92, 'Initializing annotation processor...')
    annotation_processor = AnnotationProcessor(AnnotationConfig(
        pred_output_dir=PRED_OUTPUT_DIR
    ))

    update_init_status('init_processors', 94, 'Initializing image processor...')
    image_processor = ImageProcessor(
        pdfimg_output_dir=PDFIMG_OUTPUT_DIR,
        pred_output_dir=PRED_OUTPUT_DIR
    )

    update_init_status('init_processors', 96, 'Initializing tabular processor...')
    tabular_processor = TabularProcessor(TabularConfig(
        pdfimg_output_dir=PDFIMG_OUTPUT_DIR,
        pred_output_dir=PRED_OUTPUT_DIR
    ))

    update_init_status('init_processors', 98, 'Initializing classification processor...')
    second_step_processor = SecondStepProcessor(SecondStepConfig(
        pred_output_dir=PRED_OUTPUT_DIR,
        model_path=MODELS_CLASSIFIER_DIR / "model_classifier.pth"
    ))

    update_init_status('init_processors', 99, 'Initializing export processor...')
    export_processor = ExportProcessor(ExportConfig(
        pred_output_dir=PRED_OUTPUT_DIR
    ))

    # Mark as ready
    update_init_status('ready', 100, 'Application ready!')
    init_status['ready'] = True
    print("\n" + "="*80)
    print(" ✅ All processors initialized - Application ready!")
    print("="*80)


# Initialize processor variables as None
pdf_processor = None
model_processor = None
mask_extractor = None
annotation_processor = None
image_processor = None
tabular_processor = None
second_step_processor = None
export_processor = None


def background_initialization():
    """Run initialization in background thread"""
    import threading
    
    def init_thread():
        try:
            update_init_status('init_start', 5, 'Starting initialization...')
            initialize_models()
            initialize_processors()
        except Exception as e:
            print(f"ERROR during initialization: {e}")
            import traceback
            traceback.print_exc()
            update_init_status('error', 0, f'Initialization failed: {e}')
    
    thread = threading.Thread(target=init_thread, daemon=True)
    thread.start()
    print("🚀 Background initialization started...")


# Start background initialization
background_initialization()


# ==================== ROUTES ====================

@app.route('/api/init-status')
def get_init_status():
    """Get initialization status"""
    return jsonify(init_status)

@app.route('/api/system-info')
def get_system_info():
    """Get system information including CPU, GPU, and MPS availability"""
    try:
        import os
        import torch
        
        system_info = {
            'cpu': {
                'cores': os.cpu_count() or 1,
                'available': True
            },
            'gpu': {
                'cuda_available': torch.cuda.is_available(),
                'cuda_version': torch.version.cuda if torch.cuda.is_available() else None,
                'gpu_count': torch.cuda.device_count() if torch.cuda.is_available() else 0,
                'gpu_names': [torch.cuda.get_device_name(i) for i in range(torch.cuda.device_count())] if torch.cuda.is_available() else []
            },
            'mps': {
                'mps_available': torch.backends.mps.is_available() if hasattr(torch.backends, 'mps') else False
            }
        }
        
        return jsonify(system_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

# ============================================================================
# PROJECT MANAGEMENT API ROUTES
# ============================================================================

@app.route('/api/projects', methods=['GET'])
def list_projects():
    """Get list of all projects"""
    try:
        projects = project_manager.list_projects()
        return jsonify({'projects': projects, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects', methods=['POST'])
def create_project():
    """Create a new project"""
    try:
        data = request.get_json()
        project_name = data.get('project_name', '').strip()
        description = data.get('description', '').strip()
        icon = data.get('icon', '1.png')
        
        if not project_name:
            return jsonify({'error': 'Project name is required', 'success': False}), 400
        
        metadata = project_manager.create_project(project_name, description, icon)
        return jsonify({'project': metadata, 'success': True})
    except ValueError as e:
        return jsonify({'error': str(e), 'success': False}), 400
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>', methods=['GET'])
def get_project(project_id):
    """Get project metadata"""
    try:
        metadata = project_manager.get_project(project_id)
        if metadata is None:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        return jsonify({'project': metadata, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project"""
    try:
        success = project_manager.delete_project(project_id)
        if not success:
            return jsonify({'error': 'Project not found or could not be deleted', 'success': False}), 404
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/icons', methods=['GET'])
def get_icons():
    """Get list of available project icons"""
    try:
        icons_path = Path('static/imgs/icons')
        if not icons_path.exists():
            return jsonify({'icons': [], 'success': True})

        # Get all PNG files in static/imgs/icons folder
        icons = [f.name for f in icons_path.iterdir()
                if f.is_file() and f.suffix.lower() == '.png' and f.name != 'LogoLens.png']
        
        # Sort icons numerically if they are numbered
        try:
            icons.sort(key=lambda x: int(x.replace('.png', '')))
        except:
            icons.sort()
        
        return jsonify({'icons': icons, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/icons/<filename>')
def serve_icon(filename):
    """Serve an icon file"""
    try:
        icons_path = Path('static/imgs/icons')
        return send_from_directory(icons_path, filename)
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


@app.route('/api/projects/<project_id>/workflow', methods=['PATCH'])
def update_workflow_status(project_id):
    """Update workflow status for a project"""
    try:
        data = request.get_json()
        status_updates = data.get('status_updates', {})
        
        success = project_manager.update_workflow_status(project_id, status_updates)
        if not success:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Return updated metadata
        metadata = project_manager.get_project(project_id)
        return jsonify({'project': metadata, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/settings', methods=['PATCH'])
def update_project_settings(project_id):
    """Update project settings"""
    try:
        data = request.get_json()
        settings = data.get('settings', {})
        
        success = project_manager.update_settings(project_id, settings)
        if not success:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Return updated metadata
        metadata = project_manager.get_project(project_id)
        return jsonify({'project': metadata, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/excluded-images', methods=['POST'])
def update_excluded_images(project_id):
    """Update excluded images list for a project"""
    try:
        data = request.get_json()
        excluded_images = data.get('excluded_images', [])
        
        success = project_manager.update_excluded_images(project_id, excluded_images)
        if not success:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/reviewed', methods=['POST'])
def add_reviewed_image(project_id):
    """Mark an image as reviewed"""
    try:
        data = request.get_json()
        image_name = data.get('image_name', '')
        
        if not image_name:
            return jsonify({'error': 'Image name is required', 'success': False}), 400
        
        success = project_manager.add_reviewed_image(project_id, image_name)
        if not success:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/images', methods=['GET'])
def get_project_images(project_id):
    """Get list of images in project"""
    try:
        images = project_manager.get_images_list(project_id, 'images')
        
        if images is None:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Create URLs for each image
        image_urls = [f'/api/projects/{project_id}/image/{img}' for img in images]
        
        return jsonify({
            'images': image_urls,
            'count': len(images),
            'success': True
        })
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/image/<filename>', methods=['GET'])
def serve_project_image(project_id, filename):
    """Serve an image from project's images folder"""
    try:
        images_path = project_manager.get_project_path(project_id, 'images')
        
        if not images_path or not images_path.exists():
            return jsonify({'error': 'Project or images folder not found', 'success': False}), 404
        
        return send_from_directory(images_path, filename)
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


@app.route('/api/projects/<project_id>/masks', methods=['GET'])
def get_project_masks(project_id):
    """Get list of mask images in project"""
    try:
        masks = project_manager.get_images_list(project_id, 'masks')
        
        if masks is None:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Create URLs for each mask
        mask_urls = [f'/api/projects/{project_id}/mask/{img}' for img in masks]
        
        return jsonify({
            'masks': mask_urls,
            'count': len(masks),
            'success': True
        })
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/mask/<filename>', methods=['GET'])
def serve_project_mask(project_id, filename):
    """Serve a mask image from project's masks folder"""
    try:
        masks_path = project_manager.get_project_path(project_id, 'masks')
        
        if not masks_path or not masks_path.exists():
            return jsonify({'error': 'Project or masks folder not found', 'success': False}), 404
        
        return send_from_directory(masks_path, filename)
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


@app.route('/api/projects/<project_id>/masks/extract', methods=['POST'])
def extract_project_masks(project_id):
    """Extract cards from masks in project with progress tracking"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Get project paths
        masks_path = project_manager.get_project_path(project_id, 'masks')
        cards_path = project_manager.get_project_path(project_id, 'cards')
        
        if not masks_path or not masks_path.exists():
            return jsonify({'error': 'Project masks folder not found', 'success': False}), 404
        
        # Count mask files for progress
        mask_files = [f for f in masks_path.iterdir() if f.name.endswith('_mask_layer.png')]
        total_masks = len(mask_files)
        
        if total_masks == 0:
            return jsonify({'error': 'No mask files found. Apply a model first.', 'success': False}), 404
        
        # Initialize progress
        update_operation_progress('extract_masks', 0, total_masks, 'Starting extraction...')
        
        # Run extraction in a way that allows progress updates
        # We'll monkey-patch the print function temporarily
        import builtins
        original_print = builtins.print
        
        def progress_print(*args, **kwargs):
            msg = ' '.join(str(arg) for arg in args)
            # Look for progress pattern "Processing mask X/Y"
            if 'Processing mask' in msg:
                try:
                    parts = msg.split()
                    idx = parts.index('mask') + 1
                    current = int(parts[idx].split('/')[0])
                    update_operation_progress('extract_masks', current, total_masks, 
                                             f'Extracting mask {current}/{total_masks}')
                except:
                    pass
            original_print(*args, **kwargs)
        
        builtins.print = progress_print
        
        try:
            # Extract masks using project paths
            result = mask_extractor.extract_masks_from_project(
                str(masks_path),
                str(cards_path)
            )
        finally:
            builtins.print = original_print
            clear_operation_progress()
        
        # Update project workflow status
        card_count = len(list(cards_path.glob('*.png'))) if cards_path.exists() else 0
        project_manager.update_workflow_status(project_id, {
            'cards_extracted': card_count
        })
        
        return jsonify({
            'message': result,
            'success': True
        })
        
    except Exception as e:
        clear_operation_progress()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/metadata/extract', methods=['POST'])
def extract_project_metadata(project_id):
    """Extract metadata (captions, figure numbers, pottery IDs) from project"""
    try:
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json or {}
        reference_pdf_path = data.get('reference_pdf_path')  # Optional: PDF with period info
        pdf_path = data.get('pdf_path')  # Source PDF for text extraction

        project_path = project_manager.get_project_path(project_id)
        config = MetadataExtractionConfig(project_path=project_path)
        extractor = MetadataExtractor(config)

        # Try to find source PDF if not provided
        if not pdf_path:
            # Look in pdf_source folder first
            pdf_source_path = project_manager.get_project_path(project_id, 'pdf_source')
            if pdf_source_path and pdf_source_path.exists():
                pdf_files = list(pdf_source_path.glob('*.pdf'))
                if pdf_files:
                    pdf_path = str(pdf_files[0])
                    print(f"Found PDF in pdf_source: {pdf_path}")
            # Fallback to project root folder
            if not pdf_path:
                pdf_files = list(project_path.glob('*.pdf'))
                if pdf_files:
                    pdf_path = str(pdf_files[0])
                    print(f"Found PDF in project root: {pdf_path}")

        # Extract period mappings from reference PDF if provided
        period_mappings = {}
        if reference_pdf_path:
            ref_path = Path(reference_pdf_path)
            if ref_path.exists():
                print(f"Extracting period mappings from: {ref_path}")
                period_mappings = extractor.extract_period_mappings_from_pdf(ref_path)

        result = extractor.process_project(project_id, project_manager, period_mappings, pdf_path=Path(pdf_path) if pdf_path else None)

        return jsonify({'message': result, 'success': True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/metadata/extract-periods', methods=['POST'])
def extract_period_mappings(project_id):
    """Extract pottery->period mappings from a reference PDF"""
    try:
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json
        reference_pdf_path = data.get('reference_pdf_path')

        if not reference_pdf_path:
            return jsonify({'error': 'reference_pdf_path is required', 'success': False}), 400

        ref_path = Path(reference_pdf_path)
        if not ref_path.exists():
            return jsonify({'error': f'PDF not found: {reference_pdf_path}', 'success': False}), 404

        project_path = project_manager.get_project_path(project_id)
        config = MetadataExtractionConfig(project_path=project_path)
        extractor = MetadataExtractor(config)

        mappings = extractor.extract_period_mappings_from_pdf(ref_path)

        # Save mappings to project
        mappings_path = project_path / 'period_mappings.json'
        import json
        with open(mappings_path, 'w', encoding='utf-8') as f:
            json.dump(mappings, f, indent=2, ensure_ascii=False)

        return jsonify({
            'message': f'Extracted {len(mappings)} period mappings',
            'mappings': mappings,
            'success': True
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/masks/save', methods=['POST'])
def save_project_mask(project_id):
    """Save edited mask for a project image"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Get uploaded mask file
        if 'mask' not in request.files:
            return jsonify({'error': 'No mask file provided', 'success': False}), 400
        
        mask_file = request.files['mask']
        if mask_file.filename == '':
            return jsonify({'error': 'Empty filename', 'success': False}), 400
        
        # Save to project masks folder
        masks_path = project_manager.get_project_path(project_id, 'masks')
        filename = secure_filename(mask_file.filename)
        mask_filepath = masks_path / filename
        
        mask_file.save(mask_filepath)
        
        # Generate mask URL for frontend
        mask_url = f'/api/projects/{project_id}/mask/{filename}'
        
        return jsonify({
            'message': f'Mask saved: {filename}',
            'filename': filename,
            'mask_url': mask_url,
            'success': True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


# ============================================================================
# LEGACY API ROUTES (will be refactored to use projects)
# ============================================================================

@app.route('/api/folders/images')
def get_image_folders():
    """Get list of image folders"""
    try:
        folders = [f for f in os.listdir(PDFIMG_OUTPUT_DIR) 
                  if os.path.isdir(PDFIMG_OUTPUT_DIR / f)]
        return jsonify({'folders': folders, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/folders/masks')
def get_mask_folders():
    """Get list of folders with masks (for annotation tab)"""
    try:
        # Get folders ending with _mask
        folders = [f.replace('_mask', '') for f in os.listdir(PRED_OUTPUT_DIR) 
                  if f.endswith('_mask') and os.path.isdir(PRED_OUTPUT_DIR / f)]
        return jsonify({'folders': folders, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/folders/results')
def get_results_folders():
    """Get list of result folders"""
    try:
        folders = [f for f in os.listdir(PRED_OUTPUT_DIR) 
                  if f.endswith('_card') and os.path.isdir(PRED_OUTPUT_DIR / f)]
        return jsonify({'folders': folders, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/models')
def get_models():
    """Get list of available models"""
    try:
        models = [f for f in os.listdir(MODELS_DIR) 
                 if f.endswith('.pt')]
        return jsonify({'models': models, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== PDF PROCESSING ====================

@app.route('/api/pdf/upload', methods=['POST'])
def upload_pdf():
    """Upload and process PDF file into a project"""
    try:
        print("PDF upload request received")
        print("Files:", request.files)
        print("Form data:", request.form)
        
        if 'file' not in request.files:
            print("Error: No file in request")
            return jsonify({'error': 'No file provided', 'success': False}), 400
        
        file = request.files['file']
        split_pages = request.form.get('split_pages', 'false').lower() == 'true'
        project_id = request.form.get('project_id', '').strip()
        
        print(f"File name: {file.filename}")
        print(f"Split pages: {split_pages}")
        print(f"Project ID: {project_id}")
        
        if not project_id:
            return jsonify({'error': 'No project selected', 'success': False}), 400
        
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        if file.filename == '':
            print("Error: Empty filename")
            return jsonify({'error': 'No file selected', 'success': False}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            print("Error: Not a PDF file")
            return jsonify({'error': 'Only PDF files are allowed', 'success': False}), 400
        
        # Save PDF to project's pdf_source folder
        filename = secure_filename(file.filename)
        pdf_source_path = project_manager.get_project_path(project_id, 'pdf_source')
        pdf_filepath = pdf_source_path / filename
        print(f"Saving PDF to project: {pdf_filepath}")
        file.save(pdf_filepath)
        
        # Get project images folder for output
        images_output_path = project_manager.get_project_path(project_id, 'images')
        
        # Process PDF and save images to project (use project name for image naming)
        print(f"Processing PDF to: {images_output_path}")
        result = pdf_processor.process_pdf_to_folder(
            str(pdf_filepath), 
            str(images_output_path), 
            split_pages,
            project_name=project_metadata.get('project_name', 'project')
        )
        print(f"Processing result: {result}")
        
        # Update project metadata
        image_count = project_manager.count_files(project_id, 'images')
        project_manager.update_workflow_status(project_id, {
            'pdf_processed': True,
            'pdf_count': len(list(pdf_source_path.glob('*.pdf'))),
            'images_extracted': image_count,
            'total_images': image_count
        })
        
        return jsonify({
            'message': f'PDF processed successfully. {image_count} images extracted.',
            'images_count': image_count,
            'success': True
        })
        
    except Exception as e:
        print(f"Error in PDF upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== MODEL APPLICATION ====================

@app.route('/api/model/apply', methods=['POST'])
def apply_model():
    """Apply model to images in a project"""
    try:
        data = request.json
        project_id = data.get('project_id')
        model = data.get('model')
        confidence = float(data.get('confidence', 0.5))
        diagnostic = data.get('diagnostic', False)
        kernel_size = int(data.get('kernel_size', 2))
        iterations = int(data.get('iterations', 10))
        excluded_images = data.get('excluded_images', [])
        
        if not project_id or not model:
            return jsonify({'error': 'Project and model are required', 'success': False}), 400
        
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Get project paths
        images_path = project_manager.get_project_path(project_id, 'images')
        masks_path = project_manager.get_project_path(project_id, 'masks')
        
        if not images_path or not images_path.exists():
            return jsonify({'error': 'Project images folder not found', 'success': False}), 404
        
        print(f"Applying model to project {project_id} with excluded_images: {excluded_images}")
        
        # Reset progress
        global model_progress
        model_progress = {
            'total': 0,
            'current': 0,
            'message': 'Starting...',
            'active': True
        }
        
        # Progress callback
        def update_progress(current, total, message):
            global model_progress
            model_progress['current'] = current
            model_progress['total'] = total
            model_progress['message'] = message
        
        # Run model in background thread
        def run_model():
            global model_progress
            try:
                result = model_processor.apply_model_to_project(
                    str(images_path),
                    str(masks_path),
                    model,
                    confidence,
                    diagnostic,
                    kernel_size,
                    iterations,
                    excluded_images,
                    progress_callback=update_progress
                )
                
                # Update project workflow status
                mask_count = project_manager.count_files(project_id, 'masks')
                project_manager.update_workflow_status(project_id, {
                    'model_applied': True,
                    'masks_extracted': mask_count
                })
                
                model_progress['message'] = result
                model_progress['active'] = False
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                model_progress['message'] = f'Error: {str(e)}'
                model_progress['active'] = False
        
        thread = threading.Thread(target=run_model)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'message': 'Model processing started',
            'success': True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/model/progress')
def get_model_progress():
    """Get current model processing progress"""
    global model_progress
    return jsonify(model_progress)


@app.route('/api/images/<folder>')
def get_folder_images(folder):
    """Get images in folder"""
    try:
        folder_path = PDFIMG_OUTPUT_DIR / folder
        if not folder_path.exists():
            return jsonify({'error': 'Folder not found', 'success': False}), 404
        
        images = [f for f in os.listdir(folder_path) 
                 if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        
        # Return image paths
        image_urls = [f'/api/image/{folder}/{img}' for img in images[:20]]  # Limit to 20 for preview
        
        return jsonify({
            'images': image_urls,
            'count': len(images),
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/image/<folder>/<filename>')
def get_image(folder, filename):
    """Serve image file"""
    try:
        folder_path = PDFIMG_OUTPUT_DIR / folder
        return send_from_directory(folder_path, filename)
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


# ==================== ANNOTATION ====================

@app.route('/api/annotation/images/<folder>')
def get_annotation_images(folder):
    """Get list of images for annotation"""
    try:
        folder_path = PDFIMG_OUTPUT_DIR / folder
        if not folder_path.exists():
            return jsonify({'error': 'Folder not found', 'success': False}), 404
        
        images = sorted([f for f in os.listdir(folder_path) 
                        if f.lower().endswith(('.png', '.jpg', '.jpeg'))])
        
        return jsonify({
            'images': images,
            'total': len(images),
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/annotation/load', methods=['POST'])
def load_annotation():
    """Load image and existing annotation"""
    try:
        data = request.json
        folder = data.get('folder')
        image_name = data.get('image')
        
        if not folder or not image_name:
            return jsonify({'error': 'Folder and image are required', 'success': False}), 400
        
        image_path = PDFIMG_OUTPUT_DIR / folder / image_name
        if not image_path.exists():
            return jsonify({'error': 'Image not found', 'success': False}), 404
        
        # Load annotation data
        editor_data = annotation_processor.file_selection(str(image_path))
        
        # Convert to base64 for sending to client
        from PIL import Image
        import io
        
        img = Image.fromarray(editor_data['background'])
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Load mask if exists
        mask_base64 = None
        if editor_data['layers']:
            mask_img = Image.fromarray(editor_data['layers'][0])
            mask_buffer = io.BytesIO()
            mask_img.save(mask_buffer, format='PNG')
            mask_base64 = base64.b64encode(mask_buffer.getvalue()).decode()
        
        return jsonify({
            'image': f'data:image/png;base64,{img_base64}',
            'mask': f'data:image/png;base64,{mask_base64}' if mask_base64 else None,
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/annotation/save', methods=['POST'])
def save_annotation():
    """Save annotation mask"""
    try:
        data = request.json
        folder = data.get('folder')
        image_name = data.get('image')
        mask_data = data.get('mask')  # Base64 encoded mask
        
        if not all([folder, image_name, mask_data]):
            return jsonify({'error': 'Missing required data', 'success': False}), 400
        
        # Decode mask from base64
        from PIL import Image
        import io
        import numpy as np
        
        mask_bytes = base64.b64decode(mask_data.split(',')[1])
        mask_img = Image.open(io.BytesIO(mask_bytes))
        mask_array = np.array(mask_img)
        
        # Prepare editor_data format
        editor_data = {
            'layers': [mask_array]
        }
        
        # Save
        success = annotation_processor.save_annotation(folder, editor_data, image_name)
        
        return jsonify({
            'message': 'Mask saved successfully' if success else 'Failed to save mask',
            'success': success
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/annotation/extract', methods=['POST'])
def extract_masks():
    """Extract masks from annotations"""
    try:
        data = request.json
        folder = data.get('folder')
        
        if not folder:
            return jsonify({'error': 'Folder is required', 'success': False}), 400
        
        result = mask_extractor.extract_masks(folder)
        
        return jsonify({
            'message': result,
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== TABULAR ====================

@app.route('/api/tabular/load', methods=['POST'])
def load_tabular_data():
    """Load tabular data for an image"""
    try:
        data = request.json
        folder = data.get('folder')
        img_num = int(data.get('img_num', 0))
        
        if not folder:
            return jsonify({'error': 'Folder is required', 'success': False}), 400
        
        # Get image and table data
        image_data, current_num, table_df, max_imgs = tabular_processor.image_selection(folder, img_num)
        
        # Convert image to base64 if exists
        img_base64 = None
        annotations = []
        if image_data and hasattr(image_data, 'value'):
            # Handle AnnotatedImage value
            img_array, annot_list = image_data.value
            from PIL import Image
            import io
            
            img = Image.fromarray(img_array)
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            annotations = annot_list
        
        # Convert DataFrame to dict
        table_data = table_df.to_dict('records') if not table_df.empty else []
        
        return jsonify({
            'image': f'data:image/png;base64,{img_base64}' if img_base64 else None,
            'annotations': annotations,
            'table': table_data,
            'columns': list(table_df.columns) if not table_df.empty else [],
            'current': current_num,
            'total': max_imgs,
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/tabular/save', methods=['POST'])
def save_tabular_data():
    """Save tabular data"""
    try:
        data = request.json
        folder = data.get('folder')
        table_data = data.get('table')
        
        if not folder or not table_data:
            return jsonify({'error': 'Missing required data', 'success': False}), 400
        
        # Convert to DataFrame
        df = pd.DataFrame(table_data)
        
        # Save
        tabular_processor.save_table(df, folder)
        
        return jsonify({
            'message': 'Table saved successfully',
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/tabular/add-column', methods=['POST'])
def add_column():
    """Add a new column to the table"""
    try:
        data = request.json
        column_name = data.get('column_name')
        table_data = data.get('table')
        
        if not column_name:
            return jsonify({'error': 'Column name is required', 'success': False}), 400
        
        df = pd.DataFrame(table_data)
        if column_name not in df.columns:
            df[column_name] = ""
        
        return jsonify({
            'table': df.to_dict('records'),
            'columns': list(df.columns),
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards')
def get_project_cards(project_id):
    """Get list of card images for a project"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Get cards path
        cards_path = project_manager.get_project_path(project_id, 'cards')
        
        if not cards_path or not cards_path.exists():
            return jsonify({
                'cards': [],
                'total': 0,
                'success': True
            })
        
        # Get all card images
        card_extensions = {'.jpg', '.jpeg', '.png', '.bmp'}
        cards = sorted([f.name for f in cards_path.iterdir() 
                       if f.is_file() and f.suffix.lower() in card_extensions])
        
        # Load classifications if available (check both cards and cards_modified)
        classifications = {}
        
        # Try cards_modified first (where classifications.csv is usually saved after processing)
        project_base_path = project_manager.get_project_path(project_id, 'cards')
        if project_base_path:
            project_root = project_base_path.parent
            cards_modified_path = project_root / 'cards_modified'
            classifications_csv = cards_modified_path / 'classifications.csv'
            
            if not classifications_csv.exists():
                # Fallback to cards folder
                classifications_csv = project_base_path / 'classifications.csv'
        else:
            classifications_csv = None
        
        if classifications_csv and classifications_csv.exists():
            try:
                import pandas as pd
                df = pd.read_csv(classifications_csv)
                print(f"Loaded classifications from {classifications_csv}, columns: {df.columns.tolist()}")
                
                # Create mapping from filename to type with normalization
                for _, row in df.iterrows():
                    # Try different column names for filename
                    filename = row.get('filename') or row.get('Filename') or row.get('mask_file') or row.get('id')
                    type_val = row.get('type') or row.get('Type')
                    
                    if filename and type_val:
                        # Normalize filename by removing path and keeping just the name
                        from pathlib import Path
                        filename_clean = Path(filename).name
                        classifications[filename_clean] = type_val
                        print(f"Mapped {filename_clean} -> {type_val}")
            except Exception as e:
                print(f"Error loading classifications: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"No classifications.csv found in cards or cards_modified")
        
        # Create URLs and metadata for cards
        card_data = []
        for card in cards:
            card_type = classifications.get(card, 'ENT')  # Default to ENT if not classified
            print(f"Card {card} -> type {card_type}")
            card_data.append({
                'url': f'/api/projects/{project_id}/card/{card}',
                'filename': card,
                'type': card_type
            })
        
        return jsonify({
            'cards': card_data,
            'total': len(card_data),
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/card/<filename>')
def serve_project_card(project_id, filename):
    """Serve a specific card image from project"""
    try:
        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404
        
        return send_from_directory(cards_path, filename)
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


@app.route('/api/projects/<project_id>/card-modified/<filename>')
def serve_project_card_modified(project_id, filename):
    """Serve a specific modified card image from project"""
    try:
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')
        if not cards_modified_path or not cards_modified_path.exists():
            return jsonify({'error': 'Modified cards folder not found', 'success': False}), 404
        
        return send_from_directory(cards_modified_path, filename)
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 404


@app.route('/api/projects/<project_id>/tabular/load', methods=['POST'])
def load_project_tabular_data(project_id):
    """Load tabular data for a project - shows original image with bounding boxes"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        img_num = int(data.get('img_num', 0))
        
        # Get project paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        images_path = project_manager.get_project_path(project_id, 'images')
        
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards found in project', 'success': False}), 404
        
        if not images_path or not images_path.exists():
            return jsonify({'error': 'No images found in project', 'success': False}), 404
        
        # Load annotation CSVs
        mask_info_path = cards_path / 'mask_info.csv'
        mask_info_annots_path = cards_path / 'mask_info_annots.csv'
        
        if not mask_info_path.exists() or not mask_info_annots_path.exists():
            return jsonify({'error': 'Annotation CSV files not found', 'success': False}), 404
        
        # Read CSV files
        df_info = pd.read_csv(mask_info_path).fillna('')
        df_annots = pd.read_csv(mask_info_annots_path)
        
        # Add image_name and ID columns to annotations
        df_annots['image_name'] = df_annots['mask_file'].apply(
            lambda x: x.split('_mask_layer_')[0] if isinstance(x, str) and '_mask_layer_' in x else '')
        df_annots['ID'] = df_annots['mask_file'].apply(
            lambda x: x.split('layer_')[1].replace('.png', '') if isinstance(x, str) and 'layer_' in x else '0')
        
        # Get list of unique images
        unique_images = sorted(df_annots['image_name'].unique())
        
        if not unique_images:
            return jsonify({'error': 'No images found in annotations', 'success': False}), 404
        
        # Validate image number
        img_num = max(0, min(img_num, len(unique_images) - 1))
        current_image_name = unique_images[img_num]
        
        # Find original image file
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
        original_image_path = None
        
        for ext in image_extensions:
            candidate = images_path / f"{current_image_name}{ext}"
            if candidate.exists():
                original_image_path = candidate
                break
        
        if not original_image_path:
            return jsonify({'error': f'Original image not found: {current_image_name}', 'success': False}), 404
        
        # Load and prepare image with annotations
        from PIL import Image
        import io

        with Image.open(original_image_path) as img:
            # Keep original size for bbox math
            original_size = img.size  # (width, height)

            # Create a reasonably sized preview but preserve aspect ratio (no rotation)
            img.thumbnail((1200, 1200))

            # Compute scale factors
            scale_x = img.size[0] / original_size[0]
            scale_y = img.size[1] / original_size[1]

            image_array = np.asarray(img, dtype=np.uint8)
        
        # Get annotations for this image
        image_annots = df_annots[df_annots['image_name'] == current_image_name]
        
        # Create scaled annotations list
        annotations = []
        for _, row in image_annots.iterrows():
            try:
                # Parse bbox string "(x1, y1, x2, y2)"
                bbox_str = str(row.get('bbox', '')).strip('()')
                coords = [int(x.strip()) for x in bbox_str.split(',')]

                # Scale coordinates (no rotation - keep original orientation)
                scaled_bbox = [
                    int(coords[0] * scale_x),
                    int(coords[1] * scale_y),
                    int(coords[2] * scale_x),
                    int(coords[3] * scale_y)
                ]

                annotations.append({
                    'bbox': scaled_bbox,
                    'label': str(row.get('ID', ''))
                })

            except Exception as e:
                print(f"Error processing annotation: {e}")
                continue
        
        # Convert image to base64
        buffer = io.BytesIO()
        Image.fromarray(image_array).save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
    # Prepare table data from mask_info
        df_subset = df_info[df_info['file'] == current_image_name].copy()

        if not df_subset.empty:
            # Add ID column
            df_subset['ID'] = df_subset['mask_file'].apply(
                lambda x: x.split('layer_')[1] if isinstance(x, str) and 'layer_' in x else '0')

            # Drop internal columns
            drop_cols = [col for col in ['mask_file', 'file'] if col in df_subset.columns]
            if drop_cols:
                df_subset = df_subset.drop(columns=drop_cols)

            # Check for AI extraction results and merge
            exports_path = project_manager.get_project_path(project_id, 'exports')
            if exports_path:
                ai_csv_path = exports_path / 'card_data.csv'
                if ai_csv_path.exists() and ai_csv_path.stat().st_size > 10:
                    try:
                        df_ai = pd.read_csv(ai_csv_path).fillna('')
                        print(f"[Tabular] Loaded AI data with {len(df_ai)} rows, columns: {df_ai.columns.tolist()}")
                        # Match AI data to current image's cards
                        ai_matches_found = 0
                        for idx, row in df_subset.iterrows():
                            mask_id = str(row.get('ID', '')).replace('.png', '').replace('.jpg', '')
                            # Try multiple matching patterns
                            patterns = [
                                f"{current_image_name}_mask_layer_{mask_id}.png",
                                f"{current_image_name}_mask_layer_{mask_id}.jpg",
                                f"{current_image_name}_layer_{mask_id}.png",
                                f"{current_image_name}_{mask_id}.png",
                            ]
                            ai_match = pd.DataFrame()
                            for pattern in patterns:
                                ai_match = df_ai[df_ai['filename'] == pattern]
                                if not ai_match.empty:
                                    break
                            # If no exact match, try partial match
                            if ai_match.empty and 'filename' in df_ai.columns:
                                search_term = f"layer_{mask_id}"
                                matches = df_ai[df_ai['filename'].str.contains(current_image_name, na=False) &
                                               df_ai['filename'].str.contains(search_term, na=False)]
                                if not matches.empty:
                                    ai_match = matches.head(1)

                            if not ai_match.empty:
                                ai_matches_found += 1
                                ai_row = ai_match.iloc[0]
                                # Add AI columns to the row
                                if ai_row.get('ai_figure_number'):
                                    df_subset.at[idx, 'AI Fig#'] = ai_row['ai_figure_number']
                                if ai_row.get('ai_pottery_id'):
                                    df_subset.at[idx, 'AI Pottery ID'] = ai_row['ai_pottery_id']
                                if ai_row.get('ai_period'):
                                    df_subset.at[idx, 'AI Period'] = ai_row['ai_period']
                                if ai_row.get('ai_original_period'):
                                    df_subset.at[idx, 'AI Original Period'] = ai_row['ai_original_period']
                                if ai_row.get('ai_confidence'):
                                    df_subset.at[idx, 'AI Confidence'] = f"{float(ai_row['ai_confidence']):.0%}"
                        print(f"[Tabular] Matched {ai_matches_found}/{len(df_subset)} rows with AI data for {current_image_name}")
                    except Exception as e:
                        print(f"Error loading AI data: {e}")
                        import traceback
                        traceback.print_exc()

            # Reorder with ID first
            columns_order = ['ID'] + [col for col in df_subset.columns if col != 'ID']
            df_subset = df_subset[columns_order]

            # Replace NaN with empty string to ensure valid JSON (NaN is not valid JSON)
            df_subset = df_subset.fillna('')

            table_data = df_subset.to_dict('records')
            columns = list(df_subset.columns)
        else:
            # Create empty table structure
            table_data = []
            columns = ['ID', 'Notes']
        
        # Build image list with reviewed flags
        project_meta = project_metadata
        reviewed_list = project_meta.get('workflow_status', {}).get('reviewed_images', []) if project_meta else []
        image_list = [{'image_name': name, 'reviewed': (name in reviewed_list)} for name in unique_images]

        # Prepare full-resolution image URL for zoom (let frontend fetch it when user requests zoom)
        full_image_url = None
        for ext in image_extensions:
            candidate = images_path / f"{current_image_name}{ext}"
            if candidate.exists():
                full_image_url = f'/api/projects/{project_id}/image/{candidate.name}'
                break

        return jsonify({
            'image': f'data:image/png;base64,{img_base64}',
            'annotations': annotations,
            'table': table_data,
            'columns': columns,
            'current': img_num,
            'total': len(unique_images),
            'image_name': current_image_name,
            'image_list': image_list,
            'is_reviewed': (current_image_name in reviewed_list),
            'full_image_url': full_image_url,
            'success': True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/tabular/save', methods=['POST'])
def save_project_tabular_data(project_id):
    """Save tabular data for a project - updates mask_info.csv"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        table_data = data.get('table')
        image_name = data.get('image_name')  # Current image being edited
        
        if not table_data:
            return jsonify({'error': 'Missing table data', 'success': False}), 400
        
        # Get cards folder path
        cards_path = project_manager.get_project_path(project_id, 'cards')
        
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404
        
        # Convert to DataFrame
        df_new = pd.DataFrame(table_data)
        
        # Load existing mask_info.csv
        csv_path = cards_path / 'mask_info.csv'
        
        if csv_path.exists():
            try:
                df_existing = pd.read_csv(csv_path)
                
                # Remove old data for current image
                if image_name and 'file' in df_existing.columns:
                    df_existing = df_existing[df_existing['file'] != image_name]
                
                # Add file and mask_file columns to new data if not present
                if 'file' not in df_new.columns and image_name:
                    df_new['file'] = image_name
                
                if 'mask_file' not in df_new.columns and 'ID' in df_new.columns:
                    df_new['mask_file'] = df_new['ID'].apply(
                        lambda x: f"{image_name}_mask_layer_{x}" if image_name else f"mask_layer_{x}")
                
                # Combine old and new data
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                
                # Save combined data
                df_combined.to_csv(csv_path, index=False)
                
            except Exception as e:
                print(f"Warning: Could not merge with existing CSV: {e}")
                # Just save new data
                df_new.to_csv(csv_path, index=False)
        else:
            # Add required columns if missing
            if 'file' not in df_new.columns and image_name:
                df_new['file'] = image_name
            
            if 'mask_file' not in df_new.columns and 'ID' in df_new.columns:
                df_new['mask_file'] = df_new['ID'].apply(
                    lambda x: f"{image_name}_mask_layer_{x}" if image_name else f"mask_layer_{x}")
            
            df_new.to_csv(csv_path, index=False)
        
        print(f"Saved tabular data to: {csv_path}")
        
        return jsonify({
            'message': 'Table saved successfully',
            'success': True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/tabular/export', methods=['POST'])
def export_project_tabular_csv(project_id):
    """Save combined tabular CSV (mask_info.csv) in the project folder"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404

        # Get source CSV from temp location (if exists)
        temp_csv_path = cards_path / 'mask_info.csv'
        
        # Save to project root
        project_path = project_manager.get_project_path(project_id)
        export_csv_path = project_path / f"{project_id}_mask_info.csv"
        
        if temp_csv_path.exists():
            import shutil
            shutil.copy2(temp_csv_path, export_csv_path)
        else:
            return jsonify({'error': 'Combined CSV not found. Please save tabular data first.', 'success': False}), 404

        return jsonify({
            'message': f'CSV exported to {export_csv_path.name}',
            'path': str(export_csv_path),
            'success': True
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== POST PROCESSING ====================

@app.route('/api/postprocess/process', methods=['POST'])
def process_folder():
    """Process folder with classification model"""
    try:
        data = request.json
        folder = data.get('folder')
        flip_vertical = data.get('flip_vertical', True)
        flip_horizontal = data.get('flip_horizontal', True)
        
        if not folder:
            return jsonify({'error': 'Folder is required', 'success': False}), 400
        
        # Set flip options
        second_step_processor.set_flip_options(flip_vertical, flip_horizontal)
        
        # Process
        results = second_step_processor.process_folder(folder)
        
        if results.empty:
            return jsonify({'error': 'No images were processed', 'success': False}), 400
        
        return jsonify({
            'message': f'Successfully processed {len(results)} images',
            'count': len(results),
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/postprocess/load', methods=['POST'])
def load_processed_image():
    """Load processed image for review"""
    try:
        data = request.json
        folder = data.get('folder')
        img_num = int(data.get('img_num', 0))
        
        if not folder:
            return jsonify({'error': 'Folder is required', 'success': False}), 400
        
        results = second_step_processor.load_results(folder)
        if results.empty or img_num >= len(results):
            return jsonify({'error': 'No results found', 'success': False}), 404
        
        row = results.iloc[img_num]
        
        # Load images
        from PIL import Image
        import io
        
        original_path = second_step_processor.get_original_path(folder, row['filename'])
        transformed_path = second_step_processor.get_transformed_path(folder, row['filename'])
        
        def img_to_base64(path):
            if path.exists():
                img = Image.open(path)
                buffer = io.BytesIO()
                img.save(buffer, format='PNG')
                return base64.b64encode(buffer.getvalue()).decode()
            return None
        
        original_b64 = img_to_base64(original_path)
        transformed_b64 = img_to_base64(transformed_path)
        
        return jsonify({
            'original': f'data:image/png;base64,{original_b64}' if original_b64 else None,
            'transformed': f'data:image/png;base64,{transformed_b64}' if transformed_b64 else None,
            'type': row['type'],
            'position': row['position'],
            'rotation': row['rotation'],
            'current': img_num,
            'total': len(results) - 1,
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/postprocess/flip', methods=['POST'])
def flip_image():
    """Manually flip an image"""
    try:
        data = request.json
        folder = data.get('folder')
        img_num = int(data.get('img_num', 0))
        flip_type = data.get('flip_type')  # 'vertical' or 'horizontal'
        
        if not all([folder, flip_type]):
            return jsonify({'error': 'Missing required data', 'success': False}), 400
        
        results = second_step_processor.load_results(folder)
        if results.empty or img_num >= len(results):
            return jsonify({'error': 'Image not found', 'success': False}), 404
        
        filename = results.iloc[img_num]['filename']
        
        # Flip
        flipped = second_step_processor.manual_flip(folder, filename, flip_type)
        
        if flipped is None:
            return jsonify({'error': 'Failed to flip image', 'success': False}), 500
        
        # Return updated image
        from PIL import Image
        import io
        
        buffer = io.BytesIO()
        flipped.save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        return jsonify({
            'image': f'data:image/png;base64,{img_base64}',
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/postprocess/update-type', methods=['POST'])
def update_type():
    """Update type classification"""
    try:
        data = request.json
        folder = data.get('folder')
        img_num = int(data.get('img_num', 0))
        new_type = data.get('type')
        
        if not all([folder, new_type]):
            return jsonify({'error': 'Missing required data', 'success': False}), 400
        
        results = second_step_processor.load_results(folder)
        if results.empty or img_num >= len(results):
            return jsonify({'error': 'Image not found', 'success': False}), 404
        
        filename = results.iloc[img_num]['filename']
        second_step_processor.update_result(folder, filename, {'type': new_type})
        
        return jsonify({
            'message': f'Updated type to {new_type}',
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/postprocess/merge', methods=['POST'])
def merge_annotations():
    """Merge annotations with classifications"""
    try:
        data = request.json
        folder = data.get('folder')
        
        if not folder:
            return jsonify({'error': 'Folder is required', 'success': False}), 400
        
        # Get paths
        annots_path = PRED_OUTPUT_DIR / folder / "mask_info.csv"
        transformed_folder = second_step_processor.get_transformed_folder_path(folder)
        results_path = transformed_folder / "classifications.csv"
        
        if not annots_path.exists():
            return jsonify({'error': 'Annotations file not found', 'success': False}), 404
        if not results_path.exists():
            return jsonify({'error': 'Classifications not found. Process images first.', 'success': False}), 404
        
        # Load and merge
        annots_df = pd.read_csv(annots_path)
        results_df = pd.read_csv(results_path)
        
        annots_df.rename(columns={'mask_file': 'filename'}, inplace=True)
        results_df['filename'] = results_df['filename'].str.replace('.png', '')
        
        merged_df = pd.merge(
            annots_df,
            results_df[['filename', 'type']],
            on='filename',
            how='left'
        )
        
        if 'file' in merged_df.columns:
            merged_df = merged_df.drop('file', axis=1)
        
        # Save
        output_path = transformed_folder / "merged_annotations.csv"
        merged_df.to_csv(output_path, index=False)
        
        return jsonify({
            'message': 'Successfully merged annotations with classifications',
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== EXPORT ====================

@app.route('/api/export', methods=['POST'])
def export_results():
    """Export final results"""
    try:
        data = request.json
        folder = data.get('folder')
        acronym = data.get('acronym')
        export_pdf = data.get('export_pdf', False)
        page_size = data.get('page_size', 'A4')
        scale_factor = float(data.get('scale_factor', 1.0))
        
        if not all([folder, acronym]):
            return jsonify({'error': 'Folder and acronym are required', 'success': False}), 400
        
        # Validate acronym
        if not acronym.replace('_', '').isalnum():
            return jsonify({'error': 'Acronym can only contain letters, numbers, and underscores', 'success': False}), 400
        
        result = export_processor.export_results(
            folder=folder,
            acronym=acronym,
            export_pdf=export_pdf,
            page_size=page_size,
            scale_factor=scale_factor
        )
        
        return jsonify({
            'message': result,
            'success': True
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== STATIC FILES ====================

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files"""
    return send_from_directory('static', filename)


# ==================== PROJECT-AWARE POSTPROCESSING ENDPOINTS ====================

@app.route('/api/projects/<project_id>/postprocess', methods=['POST'])
def process_project_cards(project_id):
    """Process all cards in a project with classification model"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        flip_vertical = data.get('flip_vertical', True)
        flip_horizontal = data.get('flip_horizontal', True)
        
        # Get project paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')
        
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards folder found in project', 'success': False}), 404
        
        # Create cards_modified folder
        cards_modified_path.mkdir(exist_ok=True)
        
        # Set flip options
        second_step_processor.set_flip_options(flip_vertical, flip_horizontal)
        
        # Get all card images
        card_files = sorted([f for f in cards_path.iterdir() if f.suffix.lower() in ['.png', '.jpg', '.jpeg']])
        
        if not card_files:
            return jsonify({'error': 'No card images found', 'success': False}), 404
        
        total_cards = len(card_files)
        
        # Initialize progress
        update_operation_progress('postprocess', 0, total_cards, 'Starting post-processing...')
        
        # Process each card
        results = []
        for idx, card_file in enumerate(card_files):
            try:
                # Update progress
                update_operation_progress('postprocess', idx + 1, total_cards, 
                                        f'Processing {card_file.name}')
                
                # Process the image
                type_pred, pos_pred, rot_pred, transformed_image = second_step_processor.process_image(str(card_file))
                
                if all((type_pred, pos_pred, rot_pred)) and transformed_image:
                    # Save transformed image to cards_modified
                    transformed_path = cards_modified_path / card_file.name
                    transformed_image.save(transformed_path)
                    
                    results.append({
                        'filename': card_file.name,
                        'type': type_pred,
                        'position': pos_pred,
                        'rotation': rot_pred
                    })
                    print(f"Processed {card_file.name}: Type={type_pred}, Pos={pos_pred}, Rot={rot_pred}")
                    
            except Exception as e:
                print(f"Error processing {card_file.name}: {e}")
                continue
        
        # Clear progress
        clear_operation_progress()
        
        # Save classifications
        if results:
            import pandas as pd
            results_df = pd.DataFrame(results)
            classifications_path = cards_modified_path / 'classifications.csv'
            results_df.to_csv(classifications_path, index=False)
            print(f"Saved classifications for {len(results)} cards")
        
        return jsonify({
            'message': f'Successfully processed {len(results)} cards',
            'count': len(results),
            'success': True
        })
        
    except Exception as e:
        clear_operation_progress()
        print(f"Error processing project cards: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/postprocess/flip', methods=['POST'])
def flip_project_card(project_id):
    """Flip a single card image"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        img_num = int(data.get('img_num', 0))
        flip_type = data.get('flip_type', 'vertical')
        
        # Get project paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')
        
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards folder found', 'success': False}), 404
        
        # Get list of card images
        card_files = sorted([f for f in cards_path.iterdir() if f.suffix.lower() in ['.png', '.jpg', '.jpeg']])
        
        if img_num < 0 or img_num >= len(card_files):
            return jsonify({'error': 'Invalid image number', 'success': False}), 400
        
        card_file = card_files[img_num]
        
        # Check if a modified version already exists, use that instead
        modified_file = cards_modified_path / card_file.name
        if modified_file.exists():
            source_file = modified_file
        else:
            source_file = card_file
        
        # Load image from the appropriate source
        from PIL import Image
        img = Image.open(source_file)
        
        # Apply flip
        if flip_type == 'vertical':
            img = img.transpose(Image.FLIP_TOP_BOTTOM)
        elif flip_type == 'horizontal':
            img = img.transpose(Image.FLIP_LEFT_RIGHT)
        
        # Save to cards_modified (always save the result here)
        cards_modified_path.mkdir(parents=True, exist_ok=True)
        output_path = cards_modified_path / card_file.name
        img.save(output_path)
        
        # Convert to base64 for display
        import io
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'image': f'data:image/png;base64,{img_base64}'
        })
        
    except Exception as e:
        print(f"Error flipping card: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/postprocess/update-type', methods=['POST'])
def update_project_card_type(project_id):
    """Update the type classification for a card"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        filename = data.get('filename', '')
        new_type = data.get('type', '')
        
        if not filename:
            return jsonify({'error': 'Filename is required', 'success': False}), 400
        
        # Get project cards folder
        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards folder found', 'success': False}), 404
        
        # Try cards_modified first (where classifications.csv is usually saved after processing)
        project_root = cards_path.parent
        cards_modified_path = project_root / 'cards_modified'
        classifications_csv = cards_modified_path / 'classifications.csv'
        
        if not classifications_csv.exists():
            # Fallback to cards folder
            classifications_csv = cards_path / 'classifications.csv'
        
        if not classifications_csv.exists():
            return jsonify({'error': 'Classifications file not found. Run processing first.', 'success': False}), 404
        
        # Load classifications CSV
        df = pd.read_csv(classifications_csv)
        
        # Find the row with matching filename
        mask = df['filename'] == filename
        if not mask.any():
            return jsonify({'error': f'File {filename} not found in classifications', 'success': False}), 404
        
        # Update type
        df.loc[mask, 'type'] = new_type
        
        # Save back to CSV
        df.to_csv(classifications_csv, index=False)
        
        print(f"Updated type for {filename} to {new_type} in {classifications_csv}")
        
        return jsonify({
            'success': True,
            'message': 'Type updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating type: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/postprocess/merge', methods=['POST'])
def merge_project_annotations(project_id):
    """Merge mask annotations with classifications"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        # Get project cards folder
        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards folder found', 'success': False}), 404
        
        # Check for required CSV files
        mask_info_csv = cards_path / 'mask_info.csv'
        classifications_csv = cards_path / 'classifications.csv'
        
        if not mask_info_csv.exists():
            return jsonify({'error': 'mask_info.csv not found. Extract cards first.', 'success': False}), 404
        
        if not classifications_csv.exists():
            return jsonify({'error': 'classifications.csv not found. Run processing first.', 'success': False}), 404
        
        # Load both CSVs
        df_mask = pd.read_csv(mask_info_csv)
        df_class = pd.read_csv(classifications_csv)
        
        # Merge on filename (mask_file matches image in classifications)
        merged = pd.merge(
            df_mask, 
            df_class[['image', 'type', 'is_correct']], 
            left_on='mask_file', 
            right_on='image', 
            how='left'
        )
        
        # Save merged annotations
        merged_csv = cards_path / 'merged_annotations.csv'
        merged.to_csv(merged_csv, index=False)
        
        return jsonify({
            'success': True,
            'message': f'Successfully merged {len(merged)} annotations'
        })
        
    except Exception as e:
        print(f"Error merging annotations: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards/exclude', methods=['POST'])
def toggle_card_exclusion(project_id):
    """Toggle exclusion status of a card image."""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json
        filename = data.get('filename')
        excluded = data.get('excluded', False)

        if not filename:
            return jsonify({'error': 'Filename is required', 'success': False}), 400

        # Get cards path and exclusions file
        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404

        exclusions_path = cards_path / '.exclusions.json'

        # Load existing exclusions
        exclusions = {}
        if exclusions_path.exists():
            try:
                with open(exclusions_path, 'r') as f:
                    exclusions = json.load(f)
            except json.JSONDecodeError:
                exclusions = {}

        # Update exclusion status
        if excluded:
            exclusions[filename] = True
        else:
            exclusions.pop(filename, None)

        # Save exclusions
        with open(exclusions_path, 'w') as f:
            json.dump(exclusions, f, indent=2)

        return jsonify({
            'success': True,
            'filename': filename,
            'excluded': excluded,
            'total_excluded': len([k for k, v in exclusions.items() if v])
        })

    except Exception as e:
        print(f"Error toggling exclusion: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards/exclusions', methods=['GET'])
def get_card_exclusions(project_id):
    """Get all excluded cards for a project."""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path:
            return jsonify({'excluded': [], 'success': True})

        exclusions_path = cards_path / '.exclusions.json'

        exclusions = {}
        if exclusions_path.exists():
            try:
                with open(exclusions_path, 'r') as f:
                    exclusions = json.load(f)
            except json.JSONDecodeError:
                exclusions = {}

        # Return list of excluded filenames
        excluded_list = [k for k, v in exclusions.items() if v]

        return jsonify({
            'success': True,
            'excluded': excluded_list,
            'total': len(excluded_list)
        })

    except Exception as e:
        print(f"Error getting exclusions: {e}")
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards/crop', methods=['POST'])
def crop_card(project_id):
    """Crop a card image (auto or manual)."""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json
        filename = data.get('filename')
        mode = data.get('mode', 'manual')  # 'auto', 'manual', or 'content'
        rect = data.get('rect')  # For manual mode: {x, y, width, height}
        keep_side = data.get('keep_side', 'auto')  # For auto mode: 'auto', 'left', 'right'

        if not filename:
            return jsonify({'error': 'Filename is required', 'success': False}), 400

        # Get paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')

        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404

        # Find source image (prefer modified, fallback to original)
        source_path = None
        if cards_modified_path and (cards_modified_path / filename).exists():
            source_path = cards_modified_path / filename
        elif (cards_path / filename).exists():
            source_path = cards_path / filename
        else:
            return jsonify({'error': f'Image not found: {filename}', 'success': False}), 404

        # Import crop processor and PIL
        from crop_processor import CropProcessor
        from PIL import Image

        # Load image
        img = Image.open(source_path)
        img_array = np.array(img)

        crop_processor = CropProcessor()
        metadata = {}

        if mode == 'auto':
            cropped, metadata = crop_processor.auto_remove_section(img_array, keep_side=keep_side)
        elif mode == 'content':
            cropped, metadata = crop_processor.crop_to_content(img_array)
        elif mode == 'manual':
            if not rect:
                return jsonify({'error': 'Rectangle required for manual mode', 'success': False}), 400
            cropped = crop_processor.manual_crop(img_array, rect)
            metadata = {'rect': rect}
        else:
            return jsonify({'error': f'Unknown crop mode: {mode}', 'success': False}), 400

        # Ensure cards_modified exists
        if not cards_modified_path:
            cards_modified_path = cards_path.parent / 'cards_modified'
        cards_modified_path.mkdir(parents=True, exist_ok=True)

        # Save cropped image
        output_path = cards_modified_path / filename
        Image.fromarray(cropped).save(output_path)

        # Generate base64 preview
        from io import BytesIO
        buffer = BytesIO()
        Image.fromarray(cropped).save(buffer, format='PNG')
        cropped_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return jsonify({
            'success': True,
            'filename': filename,
            'mode': mode,
            'output_path': str(output_path),
            'cropped_size': [int(cropped.shape[1]), int(cropped.shape[0])],
            'cropped_image': f'data:image/png;base64,{cropped_b64}',
            'metadata': metadata
        })

    except Exception as e:
        print(f"Error cropping image: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards/crop-preview', methods=['POST'])
def preview_crop(project_id):
    """Preview auto-crop without applying it - shows both halves."""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json
        filename = data.get('filename')

        if not filename:
            return jsonify({'error': 'Filename is required', 'success': False}), 400

        # Get paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')

        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404

        # Find source image
        source_path = None
        if cards_modified_path and (cards_modified_path / filename).exists():
            source_path = cards_modified_path / filename
        elif (cards_path / filename).exists():
            source_path = cards_path / filename
        else:
            return jsonify({'error': f'Image not found: {filename}', 'success': False}), 404

        # Import crop processor and PIL
        from crop_processor import CropProcessor
        from PIL import Image

        # Load image
        img = Image.open(source_path)
        img_array = np.array(img)

        crop_processor = CropProcessor()
        preview_data = crop_processor.auto_remove_section_preview(img_array)

        return jsonify({
            'success': True,
            'filename': filename,
            **preview_data
        })

    except Exception as e:
        print(f"Error generating crop preview: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/cards/crop-freehand', methods=['POST'])
def freehand_crop_card(project_id):
    """Crop a card image using freehand/polygon selection."""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        data = request.json
        filename = data.get('filename')
        points = data.get('points', [])  # List of [x, y] coordinates
        smoothing = data.get('smoothing', 3)

        if not filename:
            return jsonify({'error': 'Filename is required', 'success': False}), 400

        if not points or len(points) < 3:
            return jsonify({'error': 'At least 3 points are required', 'success': False}), 400

        # Convert points to tuples
        points = [(int(p[0]), int(p[1])) for p in points]

        # Get paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')

        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'Cards folder not found', 'success': False}), 404

        # Find source image
        source_path = None
        if cards_modified_path and (cards_modified_path / filename).exists():
            source_path = cards_modified_path / filename
        elif (cards_path / filename).exists():
            source_path = cards_path / filename
        else:
            return jsonify({'error': f'Image not found: {filename}', 'success': False}), 404

        # Import crop processor and PIL
        from crop_processor import CropProcessor
        from PIL import Image

        # Load image
        img = Image.open(source_path)
        img_array = np.array(img)

        crop_processor = CropProcessor()
        cropped, metadata = crop_processor.freehand_crop(img_array, points, smoothing=smoothing)

        # Ensure cards_modified exists
        if not cards_modified_path:
            cards_modified_path = cards_path.parent / 'cards_modified'
        cards_modified_path.mkdir(parents=True, exist_ok=True)

        # Save cropped image
        output_path = cards_modified_path / filename
        Image.fromarray(cropped).save(output_path)

        # Generate base64 preview
        from io import BytesIO
        buffer = BytesIO()
        Image.fromarray(cropped).save(buffer, format='PNG')
        cropped_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return jsonify({
            'success': True,
            'filename': filename,
            'mode': 'freehand',
            'output_path': str(output_path),
            'cropped_size': [int(cropped.shape[1]), int(cropped.shape[0])],
            'cropped_image': f'data:image/png;base64,{cropped_b64}',
            'metadata': metadata
        })

    except Exception as e:
        print(f"Error freehand cropping: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/projects/<project_id>/export', methods=['POST'])
def export_project_results(project_id):
    """Export final results for a project (with auto-merge if CSV exists)"""
    try:
        # Verify project exists
        project_metadata = project_manager.get_project(project_id)
        if not project_metadata:
            return jsonify({'error': 'Project not found', 'success': False}), 404
        
        data = request.json
        acronym = data.get('acronym')
        
        if not acronym:
            return jsonify({'error': 'Acronym is required', 'success': False}), 400
        
        # Validate acronym
        if not acronym.replace('_', '').isalnum():
            return jsonify({'error': 'Acronym can only contain letters, numbers, and underscores', 'success': False}), 400
        
        # Get project paths
        cards_path = project_manager.get_project_path(project_id, 'cards')
        cards_modified_path = project_manager.get_project_path(project_id, 'cards_modified')
        project_path = project_manager.get_project_path(project_id)
        exports_path = project_manager.get_project_path(project_id, 'exports')

        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards folder found', 'success': False}), 404
        
        # Auto-merge: check if combined CSV exists in project root, or use mask_info.csv from cards
        combined_csv_path = project_path / f"{project_id}_mask_info.csv"
        cards_mask_info_path = cards_path / 'mask_info.csv'

        # Prefer cards/mask_info.csv if it has metadata columns (page_num, figure_num, pottery_id, period)
        # This ensures metadata extraction results are used even if user didn't export tabular CSV
        if cards_mask_info_path.exists():
            try:
                cards_df = pd.read_csv(cards_mask_info_path)
                has_metadata = any(col in cards_df.columns for col in ['page_num', 'figure_num', 'pottery_id', 'period'])
                if has_metadata:
                    print(f"Using cards/mask_info.csv with metadata columns: {list(cards_df.columns)}")
                    # Copy to project root for consistency
                    cards_df.to_csv(combined_csv_path, index=False)
            except Exception as e:
                print(f"Warning: Could not check cards/mask_info.csv: {e}")

        if combined_csv_path.exists():
            print(f"Found combined CSV, merging with classifications...")
            try:
                # Merge the CSVs
                import pandas as pd
                
                # Load combined CSV (tabular data)
                combined_df = pd.read_csv(combined_csv_path)
                print(f"Loaded combined CSV with {len(combined_df)} rows")
                print(f"Combined CSV columns: {list(combined_df.columns)}")
                
                # Load classifications if they exist
                classifications_path = cards_modified_path / 'classifications.csv' if cards_modified_path else None
                if classifications_path and classifications_path.exists():
                    classifications_df = pd.read_csv(classifications_path)
                    print(f"Loaded classifications CSV with {len(classifications_df)} rows")
                    print(f"Classifications CSV columns: {list(classifications_df.columns)}")
                    
                    # Ensure filename columns are compatible (remove .png if present in one but not other)
                    if 'filename' in combined_df.columns and 'filename' in classifications_df.columns:
                        # Normalize filenames - remove extension for matching
                        combined_df['filename_base'] = combined_df['filename'].str.replace('.png', '').str.replace('.jpg', '')
                        classifications_df['filename_base'] = classifications_df['filename'].str.replace('.png', '').str.replace('.jpg', '')
                        
                        # Merge on normalized filename
                        merged = pd.merge(
                            combined_df,
                            classifications_df,
                            on='filename_base',
                            how='left',
                            suffixes=('', '_class')
                        )
                        
                        # Keep original filename from combined
                        if 'filename_class' in merged.columns:
                            merged = merged.drop('filename_class', axis=1)
                        merged = merged.drop('filename_base', axis=1)
                        
                        # Save merged annotations
                        cards_modified_path.mkdir(exist_ok=True)
                        merged_path = cards_modified_path / 'merged_annotations.csv'
                        merged.to_csv(merged_path, index=False)
                        print(f"Auto-merged {len(merged)} annotations to {merged_path}")
                        print(f"Merged columns: {list(merged.columns)}")
                    else:
                        print("Warning: 'filename' column not found in one of the CSVs")
                        merged_path = cards_modified_path / 'merged_annotations.csv'
                        combined_df.to_csv(merged_path, index=False)
                else:
                    # No classifications, use combined CSV directly as merged
                    cards_modified_path.mkdir(exist_ok=True)
                    merged_path = cards_modified_path / 'merged_annotations.csv'
                    combined_df.to_csv(merged_path, index=False)
                    print(f"No classifications found, using combined CSV as merged")
                    
            except Exception as e:
                print(f"Warning: Auto-merge failed: {e}")
                import traceback
                traceback.print_exc()
                # Continue with export anyway
        
        # Determine export folder (prefer cards_modified if it has content)
        if cards_modified_path and cards_modified_path.exists() and any(cards_modified_path.iterdir()):
            export_folder = cards_modified_path
        else:
            export_folder = cards_path
        
        # Load merged annotations if available
        import pandas as pd
        merged_path = cards_modified_path / 'merged_annotations.csv' if cards_modified_path else None
        
        # Create final metadata with new IDs
        metadata_df = None
        if merged_path and merged_path.exists():
            metadata_df = pd.read_csv(merged_path)
            print(f"Loaded merged annotations: {len(metadata_df)} rows")
            print(f"Merged columns: {list(metadata_df.columns)}")
        elif combined_csv_path.exists():
            # Use combined CSV if no merged exists
            metadata_df = pd.read_csv(combined_csv_path)
            print(f"Loaded combined CSV: {len(metadata_df)} rows")
            print(f"Combined columns: {list(metadata_df.columns)}")
        
        # Load classifications to ensure we have type column
        classifications_df = None
        classifications_path = cards_modified_path / 'classifications.csv' if cards_modified_path else None
        if classifications_path and classifications_path.exists():
            classifications_df = pd.read_csv(classifications_path)
            print(f"Loaded classifications: {len(classifications_df)} rows")
            print(f"Classifications columns: {list(classifications_df.columns)}")
        
        # Create ZIP in temporary location
        import tempfile
        import zipfile

        temp_dir = tempfile.mkdtemp()
        zip_path = Path(temp_dir) / f"{acronym}.zip"

        # Load AI extraction data if available
        ai_df = None
        ai_csv_path = exports_path / 'card_data.csv' if exports_path else None
        if ai_csv_path and ai_csv_path.exists() and ai_csv_path.stat().st_size > 10:
            try:
                ai_df = pd.read_csv(ai_csv_path).fillna('')
                print(f"[Export] Loaded AI data with {len(ai_df)} rows")
            except Exception as e:
                print(f"[Export] Warning: Could not load AI data: {e}")

        try:
            # Get all card images sorted
            card_images = sorted([f for f in export_folder.iterdir() if f.suffix.lower() in ['.png', '.jpg', '.jpeg']])
            print(f"Found {len(card_images)} card images before exclusion filter")

            # Filter out excluded images
            exclusions_path = cards_path / '.exclusions.json'
            exclusions = {}
            if exclusions_path.exists():
                try:
                    with open(exclusions_path, 'r') as f:
                        exclusions = json.load(f)
                except json.JSONDecodeError:
                    exclusions = {}

            if exclusions:
                card_images = [f for f in card_images if not exclusions.get(f.name, False)]
                print(f"After exclusion filter: {len(card_images)} images (excluded {len(exclusions)} images)")
            
            # Prepare final metadata with new IDs
            final_metadata = []
            
            for idx, img_file in enumerate(card_images, 1):
                new_id_with_ext = f"{acronym}_{idx}{img_file.suffix}"  # Include extension
                
                # Initialize row with id
                row_data = {'id': new_id_with_ext}
                
                # Try to find matching row in metadata
                matched = False
                if metadata_df is not None:
                    # Try different column names for matching
                    for col in ['mask_file', 'filename', 'Filename', 'file']:
                        if col in metadata_df.columns:
                            # Normalize both sides for comparison (remove extensions)
                            img_base = img_file.stem  # filename without extension
                            
                            # Try exact match first
                            mask = metadata_df[col] == img_file.name
                            if not mask.any():
                                # Try without extension
                                mask = metadata_df[col].str.replace('.png', '').str.replace('.jpg', '').str.replace('.jpeg', '') == img_base
                            
                            if mask.any():
                                row = metadata_df[mask].iloc[0]
                                
                                # Copy all columns except unwanted ones
                                exclude_cols = ['mask_file', 'filename', 'Filename', 'filename_base', 'file', 'ID', 'id']
                                for metadata_col in metadata_df.columns:
                                    if metadata_col not in exclude_cols:
                                        row_data[metadata_col] = row[metadata_col]
                                
                                matched = True
                                print(f"Matched {img_file.name} via column '{col}'")
                                break
                
                # Ensure 'type' is present from classifications if available
                if classifications_df is not None and 'type' not in row_data:
                    # Try to match with classifications
                    img_base = img_file.stem
                    for col in ['filename', 'Filename']:
                        if col in classifications_df.columns:
                            mask = classifications_df[col].str.replace('.png', '').str.replace('.jpg', '').str.replace('.jpeg', '') == img_base
                            if mask.any():
                                class_row = classifications_df[mask].iloc[0]
                                if 'type' in class_row:
                                    row_data['type'] = class_row['type']
                                    print(f"Added type '{class_row['type']}' for {img_file.name}")
                                break

                # Merge AI extraction data if available
                if ai_df is not None and 'filename' in ai_df.columns:
                    ai_match = ai_df[ai_df['filename'] == img_file.name]
                    if not ai_match.empty:
                        ai_row = ai_match.iloc[0]
                        # Add AI columns and use them to fill empty original columns
                        if ai_row.get('ai_figure_number'):
                            row_data['ai_figure_num'] = ai_row['ai_figure_number']
                            # Use AI figure_num if original is empty
                            if not row_data.get('figure_num') or str(row_data.get('figure_num', '')).strip() in ['', 'nan', 'None']:
                                row_data['figure_num'] = ai_row['ai_figure_number']
                        if ai_row.get('ai_pottery_id'):
                            row_data['ai_pottery_id'] = ai_row['ai_pottery_id']
                            # Use AI pottery_id if original is empty
                            if not row_data.get('pottery_id') or str(row_data.get('pottery_id', '')).strip() in ['', 'nan', 'None']:
                                row_data['pottery_id'] = ai_row['ai_pottery_id']
                        if ai_row.get('ai_period'):
                            row_data['ai_period'] = ai_row['ai_period']
                            # Use AI period if original is empty
                            if not row_data.get('period') or str(row_data.get('period', '')).strip() in ['', 'nan', 'None']:
                                row_data['period'] = ai_row['ai_period']
                        if ai_row.get('ai_original_period'):
                            row_data['ai_original_period'] = ai_row['ai_original_period']
                        if ai_row.get('ai_confidence'):
                            row_data['ai_confidence'] = ai_row['ai_confidence']
                        print(f"[Export] Merged AI data for {img_file.name}: period={ai_row.get('ai_period', '')}")

                final_metadata.append(row_data)
            
            # Create final metadata DataFrame
            final_df = pd.DataFrame(final_metadata)
            
            # Reorder columns: important ones first, then others alphabetically
            priority_cols = ['id', 'type', 'period', 'ai_period', 'ai_original_period', 'figure_num', 'ai_figure_num', 'page_num', 'pottery_id', 'ai_pottery_id', 'ai_confidence', 'folder', 'image_path']
            cols = [c for c in priority_cols if c in final_df.columns]
            # Add remaining columns alphabetically
            remaining = sorted([col for col in final_df.columns if col not in cols])
            cols.extend(remaining)
            final_df = final_df[cols]
            
            # Save metadata to temp file
            metadata_temp_path = Path(temp_dir) / f"{acronym}_metadata.csv"
            final_df.to_csv(metadata_temp_path, index=False)
            print(f"Created final metadata with {len(final_df)} rows and columns: {list(final_df.columns)}")
            
            # Create ZIP with subfolders organized by figure_num/page_num
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add images with new names, organized by figure/page
                for idx, img_file in enumerate(card_images, 1):
                    # Get metadata for this image to determine folder and name
                    subfolder = ""
                    pottery_id_suffix = ""

                    if not final_metadata[idx-1] is None:
                        row = final_metadata[idx-1]

                        # Determine period folder (primary level)
                        period = row.get('period', '') or row.get('ai_period', '')
                        period_folder = ""
                        if period and str(period).strip() and str(period).strip().lower() not in ['nan', 'none', '']:
                            # Sanitize period for folder name
                            period_folder = str(period).strip().replace(' ', '_').replace('/', '-').replace('\\', '-')
                        else:
                            period_folder = "Unknown_Period"

                        # Determine figure folder (secondary level)
                        figure_num = row.get('figure_num', '')
                        page_num = row.get('page_num', '')
                        figure_folder = ""

                        if figure_num and str(figure_num).strip() and str(figure_num).strip().lower() not in ['nan', 'none', '']:
                            # Sanitize figure_num for folder name
                            figure_folder = str(figure_num).strip().replace(' ', '_').replace('/', '-').replace('\\', '-')
                        elif page_num and str(page_num).strip() and str(page_num) not in ['-1', 'nan', 'None']:
                            figure_folder = f"page_{page_num}"

                        # Build hierarchical subfolder: Period/Figure
                        if figure_folder:
                            subfolder = f"{period_folder}/{figure_folder}"
                        else:
                            subfolder = period_folder

                        # Add pottery_id to filename if available (check for NaN)
                        pottery_id = row.get('pottery_id', '')
                        if pottery_id and str(pottery_id).strip() and str(pottery_id).lower() not in ['nan', 'none', '']:
                            # Clean pottery_id for filename - take only first ID if multiple
                            pottery_id_str = str(pottery_id).strip()
                            first_id = pottery_id_str.split(',')[0].strip() if ',' in pottery_id_str else pottery_id_str
                            pottery_id_clean = first_id.replace(' ', '_').replace('/', '-')
                            pottery_id_suffix = f"_{pottery_id_clean}"

                    # Build new filename
                    new_name = f"{acronym}_{idx}{pottery_id_suffix}{img_file.suffix}"

                    # Build path with subfolder
                    if subfolder:
                        zip_path_in_archive = f"{subfolder}/{new_name}"
                    else:
                        zip_path_in_archive = new_name

                    zipf.write(img_file, zip_path_in_archive)
                    print(f"Added {img_file.name} as {zip_path_in_archive}")

                    # Update final_df with the new organized path
                    final_df.loc[idx-1, 'id'] = new_name
                    if subfolder:
                        final_df.loc[idx-1, 'folder'] = subfolder

                    # Add image path (relative path in ZIP - works after extraction)
                    final_df.loc[idx-1, 'image_path'] = zip_path_in_archive

                # Re-save metadata with updated paths
                final_df.to_csv(metadata_temp_path, index=False)

                # Create Excel file with clickable hyperlinks
                excel_temp_path = Path(temp_dir) / f"{acronym}_metadata.xlsx"
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Metadata"

                    # Write header
                    for col_idx, col_name in enumerate(final_df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                        ws.cell(row=1, column=col_idx).font = Font(bold=True)

                    # Write data with hyperlinks for image_path
                    image_path_col = list(final_df.columns).index('image_path') + 1 if 'image_path' in final_df.columns else None

                    for row_idx, row in enumerate(final_df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)

                            # Make image_path a clickable hyperlink
                            if col_idx == image_path_col and value and str(value).strip():
                                # Relative path works when Excel is in same folder as extracted images
                                cell.value = str(value)
                                cell.hyperlink = str(value)
                                cell.font = Font(color="0000FF", underline="single")
                            else:
                                cell.value = value if not pd.isna(value) else ""

                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[column].width = min(max_length + 2, 50)

                    wb.save(excel_temp_path)
                    print(f"Created Excel file with hyperlinks")
                except Exception as excel_err:
                    print(f"Warning: Could not create Excel file: {excel_err}")
                    excel_temp_path = None

                # Add metadata files to ZIP
                zipf.write(metadata_temp_path, f"{acronym}_metadata.csv")
                if excel_temp_path and excel_temp_path.exists():
                    zipf.write(excel_temp_path, f"{acronym}_metadata.xlsx")
                print(f"Added metadata files")
        
            # Send the ZIP file
            return send_file(
                str(zip_path),
                as_attachment=True,
                download_name=f"{acronym}.zip",
                mimetype='application/zip'
            )
            
        finally:
            # Cleanup will happen after send_file completes
            pass
        
    except Exception as e:
        print(f"Error exporting project results: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'success': False}), 500


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Resource not found', 'success': False}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Internal server error', 'success': False}), 500


@app.errorhandler(400)
def bad_request(e):
    return jsonify({'error': 'Bad request', 'success': False}), 400


@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({'error': 'Method not allowed', 'success': False}), 405


@app.errorhandler(415)
def unsupported_media_type(e):
    return jsonify({'error': 'Unsupported media type', 'success': False}), 415


@app.route('/api/projects/<project_id>/thumbnail/<filename>')
def serve_project_thumbnail(project_id, filename):
    """Serve a thumbnail version of a project image"""
    try:
        from PIL import Image
        import io
        
        # Get images path
        images_path = project_manager.get_project_path(project_id, 'images')
        if not images_path or not images_path.exists():
            return jsonify({'error': 'Images folder not found', 'success': False}), 404
        
        image_path = images_path / filename
        if not image_path.exists():
            return jsonify({'error': 'Image not found', 'success': False}), 404
        
        # Open and create thumbnail
        with Image.open(image_path) as img:
            # Convert to RGB if necessary (for JPEG compatibility)
            if img.mode in ('RGBA', 'LA', 'P'):
                img = img.convert('RGB')
            
            # Create thumbnail (max 300px on longest side, maintain aspect ratio)
            img.thumbnail((300, 300), Image.Resampling.LANCZOS)
            
            # Save to memory buffer
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85, optimize=True)
            buffer.seek(0)
            
            return send_file(buffer, mimetype='image/jpeg')
            
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ============================================================================
# Settings API Endpoints
# ============================================================================

@app.route('/api/settings', methods=['GET'])
def get_app_settings():
    """Get application settings (API keys masked for security)"""
    try:
        settings_manager = get_settings_manager()
        settings = settings_manager.get_settings()

        # Mask API keys for security - cloud providers
        masked_settings = {
            'anthropic_api_key': settings_manager.get_masked_key('anthropic'),
            'openai_api_key': settings_manager.get_masked_key('openai'),
            'gemini_api_key': settings_manager.get_masked_key('gemini'),
            'deepseek_api_key': settings_manager.get_masked_key('deepseek'),
            'has_anthropic_key': settings_manager.has_api_key('anthropic'),
            'has_openai_key': settings_manager.has_api_key('openai'),
            'has_gemini_key': settings_manager.has_api_key('gemini'),
            'has_deepseek_key': settings_manager.has_api_key('deepseek'),
            'default_ai_provider': settings.get('default_ai_provider', 'anthropic'),
            'calibration': settings.get('calibration', {}),
            # Local provider settings
            'lmstudio_base_url': settings.get('lmstudio_base_url', 'http://localhost:1234/v1'),
            'lmstudio_model': settings.get('lmstudio_model', ''),
            'ollama_base_url': settings.get('ollama_base_url', 'http://localhost:11434'),
            'ollama_model': settings.get('ollama_model', 'llava')
        }

        return jsonify({'settings': masked_settings, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/settings/api-key', methods=['POST'])
def update_api_key():
    """Update an API key"""
    try:
        data = request.json
        provider = data.get('provider')
        key = data.get('key', '').strip()

        valid_providers = ['anthropic', 'openai', 'gemini', 'deepseek']
        if provider not in valid_providers:
            return jsonify({'error': f'Invalid provider. Must be one of: {valid_providers}', 'success': False}), 400

        if not key:
            return jsonify({'error': 'API key cannot be empty', 'success': False}), 400

        settings_manager = get_settings_manager()
        success = settings_manager.set_api_key(provider, key)

        if success:
            return jsonify({
                'success': True,
                'message': f'{provider.capitalize()} API key saved successfully',
                'masked_key': settings_manager.get_masked_key(provider)
            })
        else:
            return jsonify({'error': 'Failed to save API key', 'success': False}), 500

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/settings/api-key/<provider>', methods=['DELETE'])
def delete_api_key(provider):
    """Delete an API key"""
    try:
        valid_providers = ['anthropic', 'openai', 'gemini', 'deepseek']
        if provider not in valid_providers:
            return jsonify({'error': f'Invalid provider. Must be one of: {valid_providers}', 'success': False}), 400

        settings_manager = get_settings_manager()
        success = settings_manager.delete_api_key(provider)

        if success:
            return jsonify({'success': True, 'message': f'{provider.capitalize()} API key deleted'})
        else:
            return jsonify({'error': 'Failed to delete API key', 'success': False}), 500

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/settings/local-provider', methods=['POST'])
def update_local_provider():
    """Update local provider settings (LM Studio, Ollama)"""
    try:
        data = request.json
        provider = data.get('provider')
        base_url = data.get('base_url')
        model = data.get('model')

        if provider not in ['lmstudio', 'ollama']:
            return jsonify({'error': 'Invalid local provider. Must be "lmstudio" or "ollama"', 'success': False}), 400

        settings_manager = get_settings_manager()
        success = settings_manager.set_local_provider_settings(provider, base_url=base_url, model=model)

        if success:
            return jsonify({
                'success': True,
                'message': f'{provider} settings saved successfully'
            })
        else:
            return jsonify({'error': 'Failed to save settings', 'success': False}), 500

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/settings/default-provider', methods=['POST'])
def set_default_provider():
    """Set the default AI provider"""
    try:
        data = request.json
        provider = data.get('provider')

        valid_providers = ['anthropic', 'openai', 'gemini', 'deepseek', 'lmstudio', 'ollama']
        if provider not in valid_providers:
            return jsonify({'error': f'Invalid provider. Must be one of: {valid_providers}', 'success': False}), 400

        settings_manager = get_settings_manager()
        success = settings_manager.set_default_provider(provider)

        if success:
            return jsonify({'success': True, 'message': f'Default provider set to {provider}'})
        else:
            return jsonify({'error': 'Failed to set default provider', 'success': False}), 500

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/settings/calibration', methods=['POST'])
def set_calibration():
    """Set calibration for project or specific image"""
    try:
        data = request.json
        pixels_per_cm = data.get('pixels_per_cm')
        image_name = data.get('image_name')  # None for default

        if pixels_per_cm is None or pixels_per_cm <= 0:
            return jsonify({'error': 'Invalid pixels_per_cm value', 'success': False}), 400

        settings_manager = get_settings_manager()
        success = settings_manager.set_calibration(pixels_per_cm, image_name)

        if success:
            msg = f'Calibration set for {image_name}' if image_name else 'Default calibration set'
            return jsonify({'success': True, 'message': msg})
        else:
            return jsonify({'error': 'Failed to set calibration', 'success': False}), 500

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ============================================================================
# Scale Bar Detection Endpoints
# ============================================================================

@app.route('/api/projects/<project_id>/scale-bar/detect', methods=['POST'])
def detect_scale_bar(project_id):
    """Auto-detect scale bar in an image"""
    try:
        import cv2

        data = request.json
        image_name = data.get('image_name')

        if not image_name:
            return jsonify({'error': 'No image specified', 'success': False}), 400

        # Get image path
        images_path = project_manager.get_project_path(project_id, 'images')
        if not images_path:
            return jsonify({'error': 'Project not found', 'success': False}), 404

        image_path = images_path / image_name
        if not image_path.exists():
            return jsonify({'error': 'Image not found', 'success': False}), 404

        # Detect scale bar
        image = cv2.imread(str(image_path))
        if image is None:
            return jsonify({'error': 'Failed to load image', 'success': False}), 500

        detector = ScaleBarDetector(ScaleBarConfig())
        result = detector.detect(image)

        if result.detected:
            return jsonify({
                'success': True,
                'result': {
                    'pixels': result.pixels,
                    'cm': result.cm,
                    'unit_text': result.unit_text,
                    'confidence': result.confidence,
                    'pixels_per_cm': result.pixels_per_cm
                }
            })
        else:
            return jsonify({
                'success': True,
                'result': None,
                'message': 'No scale bar detected'
            })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


# ============================================================================
# AI Metadata Extraction Endpoints
# ============================================================================

@app.route('/api/projects/<project_id>/metadata/ai-extract', methods=['POST'])
def ai_extract_metadata(project_id):
    """AI-powered metadata extraction using various AI providers - TWO-PASS approach"""
    try:
        from ai_extractor import DocumentStructureAnalyzer

        data = request.json
        provider = data.get('provider', 'anthropic')

        valid_providers = ['anthropic', 'openai', 'gemini', 'deepseek', 'lmstudio', 'ollama']
        if provider not in valid_providers:
            return jsonify({'error': f'Invalid provider. Must be one of: {valid_providers}', 'success': False}), 400

        # Get API key and settings from settings manager
        settings_manager = get_settings_manager()

        # Cloud providers need API keys
        api_key = ""
        base_url = ""
        model = ""

        if provider in ['anthropic', 'openai', 'gemini', 'deepseek']:
            api_key = settings_manager.get_api_key(provider)
            if not api_key:
                return jsonify({
                    'error': f'No {provider} API key configured. Please add your API key in the settings.',
                    'success': False
                }), 400
        elif provider in ['lmstudio', 'ollama']:
            # Local providers - get base_url and model from settings
            local_settings = settings_manager.get_local_provider_settings(provider)
            base_url = local_settings.get('base_url', '')
            model = local_settings.get('model', '')

        # Get project cards
        cards_path = project_manager.get_project_path(project_id, 'cards')
        if not cards_path or not cards_path.exists():
            return jsonify({'error': 'No cards found for this project', 'success': False}), 404

        # Get card images
        card_files = sorted([
            f for f in cards_path.iterdir()
            if f.suffix.lower() in ['.jpg', '.jpeg', '.png']
        ])

        if not card_files:
            return jsonify({'error': 'No card images found', 'success': False}), 404

        # Get PDF context - first check user-provided path, then project pdf_source
        pdf_context = ""
        pdf_path = data.get('pdf_path', '')  # User-provided PDF path from widget

        # Try user-provided path first
        if pdf_path and Path(pdf_path).exists():
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(pdf_path)
                print(f"[AI Extract] Reading PDF from user path: {pdf_path} ({len(doc)} pages)")
                for page_num in range(len(doc)):  # Read ALL pages for full context
                    pdf_context += doc[page_num].get_text()
                doc.close()
                print(f"[AI Extract] Extracted {len(pdf_context)} chars from PDF")
            except Exception as e:
                print(f"Warning: Could not read user PDF: {e}")

        # Fall back to project pdf_source if no user PDF
        if not pdf_context:
            pdf_source_path = project_manager.get_project_path(project_id, 'pdf_source')
            if pdf_source_path and pdf_source_path.exists():
                pdf_files = list(pdf_source_path.glob('*.pdf'))
                if pdf_files:
                    try:
                        import fitz  # PyMuPDF
                        doc = fitz.open(str(pdf_files[0]))
                        print(f"[AI Extract] Reading PDF from project: {pdf_files[0]} ({len(doc)} pages)")
                        for page_num in range(len(doc)):  # Read ALL pages
                            pdf_context += doc[page_num].get_text()
                        doc.close()
                        print(f"[AI Extract] Extracted {len(pdf_context)} chars from PDF")
                    except Exception as e:
                        print(f"Warning: Could not extract PDF context: {e}")

        # Initialize extractor
        extractor = get_extractor(provider, api_key=api_key, base_url=base_url, model=model)

        # ========================================================================
        # PASS 1: Analyze Document Structure
        # ========================================================================
        document_structure = None
        if pdf_context:
            update_operation_progress('ai_extract', 0, len(card_files), 'Analyzing document structure...')
            print("[AI Extract] PASS 1: Analyzing document structure")

            try:
                structure_analyzer = DocumentStructureAnalyzer(extractor)
                document_structure = structure_analyzer.analyze_document(pdf_context)

                if document_structure.analyzed:
                    print(f"[AI Extract] Document structure analyzed:")
                    print(f"  - Tafel/Period mappings: {len(document_structure.tafel_period_map)}")
                    print(f"  - Figure ranges: {len(document_structure.figure_ranges)}")
                    print(f"  - Catalog entries: {len(document_structure.catalog_entries)}")
                    print(f"  - Language: {document_structure.language}")
                    for tafel, period in list(document_structure.tafel_period_map.items())[:5]:
                        print(f"    {tafel} -> {period}")
                else:
                    print(f"[AI Extract] Document structure analysis failed: {document_structure.error}")
            except Exception as e:
                print(f"[AI Extract] Warning: Could not analyze document structure: {e}")

        # ========================================================================
        # Process cards with progress updates
        # ========================================================================
        results = []
        total = len(card_files)

        # Store page-specific context if PDF was loaded
        page_contexts = {}
        pdf_to_use = pdf_path if pdf_path and Path(pdf_path).exists() else None

        # If no user PDF, try to use project's PDF
        if not pdf_to_use:
            pdf_source_path = project_manager.get_project_path(project_id, 'pdf_source')
            if pdf_source_path and pdf_source_path.exists():
                pdf_files = list(pdf_source_path.glob('*.pdf'))
                if pdf_files:
                    pdf_to_use = str(pdf_files[0])
                    print(f"[AI Extract] Using project PDF: {pdf_to_use}")

        if pdf_to_use:
            try:
                import fitz
                doc = fitz.open(pdf_to_use)
                print(f"[AI Extract] Building page contexts from {len(doc)} pages")
                for page_num in range(len(doc)):
                    page_contexts[page_num + 1] = doc[page_num].get_text()  # 1-indexed
                doc.close()
                print(f"[AI Extract] Built {len(page_contexts)} page contexts")
            except Exception as e:
                print(f"Warning: Could not extract page contexts: {e}")

        # ========================================================================
        # PASS 2: Extract Metadata Per-Image with Structure Support
        # ========================================================================
        print("[AI Extract] PASS 2: Extracting metadata per image")

        for i, card_path in enumerate(card_files):
            update_operation_progress('ai_extract', i + 1, total, f'Processing {card_path.name}')

            try:
                image_b64 = image_to_base64(str(card_path))
                media_type = detect_image_media_type(str(card_path))

                # Extract page number from filename (e.g., "dopper_page_185_mask_layer_0.png")
                import re
                page_match = re.search(r'page_(\d+)', card_path.name)
                page_num = int(page_match.group(1)) if page_match else None

                # Build context: page-specific + surrounding pages + general context
                context_parts = []

                if page_num and page_contexts:
                    # Add context from the specific page and surrounding pages
                    for p in range(max(1, page_num - 1), min(len(page_contexts) + 1, page_num + 2)):
                        if p in page_contexts:
                            context_parts.append(f"=== PAGE {p} ===\n{page_contexts[p][:4000]}")

                # Add general PDF context if no page-specific
                if not context_parts and pdf_context:
                    context_parts.append(pdf_context[:12000])

                full_context = "\n\n".join(context_parts)

                # Pass document structure to extractor
                result = extractor.extract_metadata(
                    image_b64,
                    full_context[:16000],
                    media_type,
                    document_structure=document_structure
                )

                # Fallback: use document structure to look up period if AI didn't find one
                period = result.period
                if not period and document_structure and document_structure.analyzed:
                    looked_up_period = document_structure.lookup_period(
                        result.figure_number,
                        result.pottery_id
                    )
                    if looked_up_period:
                        period = looked_up_period
                        print(f"[AI Extract] Period lookup for {card_path.name}: {result.figure_number} -> {period}")

                results.append({
                    'card': card_path.name,
                    'success': result.success,
                    'figure_number': result.figure_number,
                    'pottery_id': result.pottery_id,
                    'period': period,
                    'original_period': result.original_period,
                    'confidence': result.confidence,
                    'error': result.error if not result.success else ''
                })

            except Exception as e:
                results.append({
                    'card': card_path.name,
                    'success': False,
                    'error': str(e)
                })

        clear_operation_progress()

        # Save results to project tabular data
        try:
            _save_ai_results_to_tabular(project_id, results)
        except Exception as e:
            print(f"Warning: Could not save AI results to tabular: {e}")

        successful = sum(1 for r in results if r.get('success', False))

        return jsonify({
            'success': True,
            'processed': total,
            'successful': successful,
            'results': results
        })

    except Exception as e:
        clear_operation_progress()
        return jsonify({'error': str(e), 'success': False}), 500


def _save_ai_results_to_tabular(project_id: str, results: list):
    """Save AI extraction results to the project's tabular data."""
    import pandas as pd

    exports_path = project_manager.get_project_path(project_id, 'exports')
    if not exports_path:
        return

    exports_path.mkdir(parents=True, exist_ok=True)
    csv_path = exports_path / 'card_data.csv'

    # Load existing data or create new
    if csv_path.exists() and csv_path.stat().st_size > 10:  # Check file has content
        try:
            df = pd.read_csv(csv_path)
        except Exception:
            df = pd.DataFrame()
    else:
        df = pd.DataFrame()

    # Update with AI results
    for result in results:
        if not result.get('success'):
            continue

        card_name = result.get('card', '')
        if not card_name:
            continue

        # Find or create row for this card
        if 'filename' in df.columns and card_name in df['filename'].values:
            idx = df[df['filename'] == card_name].index[0]
        else:
            idx = len(df)
            df.loc[idx, 'filename'] = card_name

        # Update fields (don't overwrite if AI result is empty)
        if result.get('figure_number'):
            df.loc[idx, 'ai_figure_number'] = result['figure_number']
        if result.get('pottery_id'):
            df.loc[idx, 'ai_pottery_id'] = result['pottery_id']
        if result.get('period'):
            df.loc[idx, 'ai_period'] = result['period']
        if result.get('original_period'):
            df.loc[idx, 'ai_original_period'] = result['original_period']
        if result.get('confidence'):
            df.loc[idx, 'ai_confidence'] = result['confidence']

    # Save updated data
    df.to_csv(csv_path, index=False)


# ============================================================================
# Operation Progress Endpoint
# ============================================================================

@app.route('/api/operation-progress', methods=['GET'])
def get_operation_progress():
    """Get current operation progress for frontend polling"""
    return jsonify({
        'success': True,
        'progress': operation_progress
    })


if __name__ == '__main__':
    print("\n" + "="*80)
    print(" 🏺 PyPotteryLens Flask Application 🔍")
    print("="*80)
    print("\n🚀 Starting server...")
    print("📝 Browser will open at: http://localhost:5001")
    print("💡 Initialization will continue in background...")
    print("\n" + "="*80 + "\n")
    
    # Open browser immediately after Flask starts
    import webbrowser
    import threading
    
    def open_browser():
        import time
        time.sleep(1)  # Wait 1 second for Flask to start
        webbrowser.open('http://localhost:5001')
        print("🌐 Browser opened!")
    
    # Start browser in separate thread
    threading.Thread(target=open_browser, daemon=True).start()
    
    app.run(
        host='0.0.0.0',
        port=5001,
        debug=True,
        threaded=True,
        use_reloader=False  # Disable reloader to prevent double initialization
    )
